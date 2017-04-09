import openpyxl
import numpy as np

class ExcelManager(object):

    url = None  # FileUrl
    wb = None  # ExcelBook
    ws = None  # ExcelWorkSeat
    row = None # ExcelRow
    # return list - > 0 to 8 chanel in EEG Module

    def read(self, path):
        self.wb = openpyxl.load_workbook(path, data_only=True)
        self.ws = self.wb.active
        excel = list(self.iter_cols(self.ws))
        self.row = self.ws.max_row
        result = np.zeros((8, self.row))
        for cnt in range(1, 9):
            min = cnt * self.row
            max = (cnt + 1) * self.row -1
            for cnt2 in range(0, self.row):
                result[cnt-1][cnt2] = excel[0][min]
                min += 1
        return result

        #for cnt in range(1, 9):
        #    min = cnt * row
        #    max = (cnt + 1) * row
        #    sub = l[0][min:max]
        #    result.append(sub)
        #return result

    # parse Excel Point and Value
    # return List of Values
    def iter_cols(self, ws):
        result = []
        for col in ws.iter_cols():
            for cell in col:
                result.append(cell.value)
        yield result